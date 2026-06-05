#!/usr/bin/env python3
"""
Feishu Bitable Monitor & Excel Export Tool

Monitors a Feishu Bitable (多维表格) for recently added/updated records,
supports conversational parameter selection, and exports to Excel matching
the parameter database template format.

Commands:
  fetch    - Connect to Feishu table, discover fields, fetch records with time filter
  export   - Export selected parameters to Excel template format

Usage:
  python feishu_monitor.py fetch --app-id ID --app-secret SECRET --app-token TOKEN --table-id TID [--days 7] [--output json]
  python feishu_monitor.py export --app-id ID --app-secret SECRET --app-token TOKEN --table-id TID --params "code1,code2" --output result.xlsx

Environment variables (alternative to CLI args):
  FEISHU_APP_ID, FEISHU_APP_SECRET, FEISHU_APP_TOKEN, FEISHU_TABLE_ID
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse


# ── Feishu API Helpers ──────────────────────────────────────────────

FEISHU_BASE = "https://open.feishu.cn/open-apis"

# ── Default Credentials ─────────────────────────────────────────────
# Built-in defaults for the shared public bitable. Users can override
# via CLI args or environment variables.
DEFAULT_APP_ID = "cli_a862482145fb100d"
DEFAULT_APP_SECRET = "1APrEPrDOlInseqsQXG99dF4VjD1OmVJ"
DEFAULT_APP_TOKEN = "I92awJKKJiV5efkYLmQctyHsn4e"
DEFAULT_TABLE_ID = "tbl7w2cvN5ZkuA7u"


def get_tenant_access_token(app_id: str, app_secret: str) -> str:
    """Obtain tenant_access_token from Feishu Open API."""
    url = f"{FEISHU_BASE}/auth/v3/tenant_access_token/internal"
    body = json.dumps({"app_id": app_id, "app_secret": app_secret}).encode("utf-8")
    req = Request(url, data=body, method="POST")
    req.add_header("Content-Type", "application/json; charset=utf-8")
    try:
        with urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except HTTPError as e:
        err_body = e.read().decode("utf-8") if e.fp else str(e)
        raise RuntimeError(f"Feishu auth failed (HTTP {e.code}): {err_body}")
    if data.get("code") != 0:
        raise RuntimeError(f"Feishu auth error: code={data.get('code')}, msg={data.get('msg')}")
    return data["tenant_access_token"]


def feishu_get(token: str, path: str, params: dict = None) -> dict:
    """Make a GET request to Feishu Open API."""
    url = f"{FEISHU_BASE}{path}"
    if params:
        qs_parts = []
        for k, v in params.items():
            qs_parts.append(f"{k}={v}")
        url += "?" + "&".join(qs_parts)
    req = Request(url, method="GET")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type", "application/json; charset=utf-8")
    try:
        with urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except HTTPError as e:
        err_body = e.read().decode("utf-8") if e.fp else str(e)
        raise RuntimeError(f"Feishu API error (HTTP {e.code}): {err_body}")
    if data.get("code") != 0:
        raise RuntimeError(f"Feishu API error: code={data.get('code')}, msg={data.get('msg')}")
    return data.get("data", {})


def parse_table_url(url: str) -> tuple:
    """
    Parse a Feishu bitable URL to extract app_token and table_id.
    Supports formats:
      - https://{domain}/wiki/{app_token}?table={table_id}&view={view_id}
      - https://{domain}/base/{app_token}?table={table_id}
    """
    parsed = urlparse(url)
    app_token = None
    table_id = None

    # Extract from path
    path_parts = parsed.path.strip("/").split("/")
    for part in path_parts:
        if part and not part.startswith("wiki") and not part.startswith("base"):
            # The token-like part (alphanumeric with possible hyphens)
            if len(part) >= 10:
                app_token = part
                break

    # Extract from query
    from urllib.parse import parse_qs
    qs = parse_qs(parsed.query)
    if "table" in qs:
        table_id = qs["table"][0]

    return app_token, table_id


# ── Table Operations ────────────────────────────────────────────────

def list_fields(token: str, app_token: str, table_id: str) -> list:
    """List all fields in a bitable table."""
    result = feishu_get(token, f"/bitable/v1/apps/{app_token}/tables/{table_id}/fields")
    items = result.get("items", [])

    fields = []
    for item in items:
        fields.append({
            "field_id": item.get("field_id", ""),
            "field_name": item.get("field_name", ""),
            "type": item.get("type", 0),
            "type_name": _field_type_name(item.get("type", 0)),
        })
    return fields


def list_records(token: str, app_token: str, table_id: str, page_size: int = 100) -> list:
    """List all records from a bitable table (paginated)."""
    all_records = []
    page_token = None

    while True:
        params = {"page_size": page_size}
        if page_token:
            params["page_token"] = page_token

        result = feishu_get(
            token,
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records",
            params=params,
        )

        items = result.get("items", [])
        for item in items:
            record = {
                "record_id": item.get("record_id", ""),
                "fields": item.get("fields", {}),
                "created_time": _parse_timestamp(item.get("created_time", 0)),
                "last_modified_time": _parse_timestamp(item.get("last_modified_time", 0)),
                "created_by": item.get("created_by", {}).get("name", ""),
                "last_modified_by": item.get("last_modified_by", {}).get("name", ""),
            }
            all_records.append(record)

        if not result.get("has_more", False):
            break
        page_token = result.get("page_token", "")

    return all_records


# ── Time Filtering ──────────────────────────────────────────────────

def filter_records_by_time(records: list, days: int) -> dict:
    """
    Filter records by time range and classify as new or updated.

    Returns:
        {
            "new": [...],       # records created within time range
            "updated": [...],   # records updated (but not created) within time range
            "all": [...],       # all records in range
            "since": "ISO datetime",
            "total_new": int,
            "total_updated": int,
        }
    """
    since = datetime.now(timezone.utc) - timedelta(days=days)
    new_records = []
    updated_records = []
    all_in_range = []

    for rec in records:
        created = rec.get("created_time")
        modified = rec.get("last_modified_time")

        is_new = created and created >= since
        is_updated = modified and modified >= since and not is_new

        if is_new or is_updated:
            all_in_range.append(rec)
            if is_new:
                new_records.append(rec)
            else:
                updated_records.append(rec)

    return {
        "new": new_records,
        "updated": updated_records,
        "all": sorted(all_in_range, key=lambda r: r.get("last_modified_time") or datetime.min, reverse=True),
        "since": since.isoformat(),
        "total_new": len(new_records),
        "total_updated": len(updated_records),
    }


# ── Display Helpers ─────────────────────────────────────────────────

def display_summary(fields: list, filter_result: dict, days: int):
    """Print a human-readable summary of the monitoring results."""
    print(f"\n{'='*60}")
    print(f"  飞书多维表格监控报告")
    print(f"{'='*60}")
    print(f"  监控时间范围: 近 {days} 天 (自 {filter_result['since'][:19]})")
    print(f"  表格字段数:   {len(fields)}")
    print(f"  新增记录:     {filter_result['total_new']} 条")
    print(f"  更新记录:     {filter_result['total_updated']} 条")
    print(f"  变动合计:     {len(filter_result['all'])} 条")
    print(f"{'='*60}\n")

    if not filter_result["all"]:
        print("  [无变动] 近 {days} 天内没有新增或更新的记录。")
        return

    # Print new records
    if filter_result["new"]:
        print(f"  📌 新增记录 ({filter_result['total_new']} 条):")
        for rec in filter_result["new"]:
            _print_record(rec, fields, tag="新增")

    # Print updated records
    if filter_result["updated"]:
        print(f"\n  ✏️  更新记录 ({filter_result['total_updated']} 条):")
        for rec in filter_result["updated"]:
            _print_record(rec, fields, tag="更新")


def _print_record(record: dict, fields: list, tag: str):
    """Print a single record summary."""
    fid = record.get("record_id", "")[:12]
    fvals = record.get("fields", {})

    # Try to find key identifying fields
    identifiers = []
    for field in fields:
        fname = field["field_name"]
        if fname in fvals and fvals[fname]:
            val = _format_field_value(fvals[fname])
            identifiers.append(f"{fname}={val}")

    ctime = record.get("created_time")
    mtime = record.get("last_modified_time")
    time_str = ""
    if mtime:
        time_str = mtime.strftime("%Y-%m-%d %H:%M") if hasattr(mtime, 'strftime') else str(mtime)[:16]

    print(f"    [{tag}] {time_str} | {' | '.join(identifiers[:4])}")


def _format_field_value(val) -> str:
    """Format a field value for display."""
    if val is None:
        return "(空)"
    if isinstance(val, list):
        items = []
        for v in val:
            if isinstance(v, dict):
                items.append(v.get("text", str(v)))
            else:
                items.append(str(v))
        return ", ".join(items[:3])
    if isinstance(val, dict):
        return val.get("text", json.dumps(val, ensure_ascii=False))
    return str(val)[:50]


def _field_type_name(type_code: int) -> str:
    """Convert Feishu field type code to readable name."""
    type_map = {
        1: "文本", 2: "数字", 3: "单选", 4: "多选",
        5: "日期", 7: "复选框", 11: "人员", 13: "电话号码",
        15: "超链接", 17: "附件", 18: "关联", 19: "查找引用",
        20: "公式", 21: "双向关联", 22: "位置", 23: "群组",
        1001: "创建时间", 1002: "最后更新时间", 1003: "创建人",
        1004: "最后更新人", 99: "自动编号",
    }
    return type_map.get(type_code, f"类型{type_code}")


def _parse_timestamp(ts):
    """Parse Feishu millisecond timestamp to datetime."""
    if not ts:
        return None
    try:
        ts_sec = int(ts) / 1000.0
        return datetime.fromtimestamp(ts_sec, tz=timezone.utc)
    except (ValueError, OSError):
        return None


# ── Excel Export ────────────────────────────────────────────────────

# Metadata field labels (rows) in the template, in order
METADATA_FIELDS = [
    "参数申请编码",
    "所属子领域",
    "所属应用编码",
    "关联微服务编码",
    "申请人",
    "联系方式",
    "主管领导",
    "参数审定编码",
    "参数类型",
    "组织职能类型",
    "参数值类型",
    "是否多选",
    "参照服务域",
    "参照编码",
    "是否允许控制下级",
    "参数取值范围",
    "参数值互斥条件",
    "取值报错提示信息",
    "参数整体介绍说明",
    "是否可为空",
    "是否默认隐藏",
    "是否校验类",
    "校验接口说明",
    "默认值",
    "参数值1说明",
    "参数值1效果图",
    "参数值2说明",
    "参数值2效果图",
    "参数值3说明",
    "参数值3效果图",
    "参数值4说明",
    "参数值4效果图",
    "参数值5说明",
    "参数值5效果图",
    "其他申请资料文档",
    "申请时间",
    "备注",
    "请填写参数注册绑定的微服务编码",
    "参数名称多语资源ID",
    "参数描述多语资源ID",
    "公有云上线时间",
    "私有云上线时间",
    "下线时间",
]


def export_to_excel(
    records: list,
    fields: list,
    selected_param_codes: list,
    output_path: str,
):
    """
    Export selected parameter records to Excel matching the template format.

    The template format is transposed:
      - Column A: metadata field labels
      - Columns B+: one column per parameter, with the parameter name as header
      - Each row is a metadata field
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Build a lookup: field_name -> field_id (for Feishu fields)
    field_name_to_id = {f["field_name"]: f["field_id"] for f in fields}

    # Match records to selected parameter codes
    # The "参数申请编码" field in Feishu maps to the identifying code
    selected_records = []
    for rec in records:
        fvals = rec.get("fields", {})
        # Try to find the parameter code in various possible field names
        code = (
            fvals.get("参数申请编码")
            or fvals.get("参数编码")
            or fvals.get("参数审定编码")
            or fvals.get("编码")
            or ""
        )
        code_str = _format_field_value(code).strip()
        if code_str in selected_param_codes:
            selected_records.append((code_str, rec))

    if not selected_records:
        # If no direct match, try fuzzy matching with field values
        for rec in records:
            fvals = rec.get("fields", {})
            for key, val in fvals.items():
                val_str = _format_field_value(val).strip()
                if val_str in selected_param_codes:
                    selected_records.append((val_str, rec))
                    break

    if not selected_records:
        raise ValueError(
            f"未找到匹配的参数记录。请在飞书表格中确认参数编码字段。"
            f"\n  指定的编码: {', '.join(selected_param_codes)}"
        )

    # Get parameter name field
    param_name_field = "参数名称"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Styles ──
    header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
    label_font = Font(name="微软雅黑", size=10, bold=False, color="000000")
    value_font = Font(name="微软雅黑", size=10, bold=False, color="000000")
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Header Row (Row 1) ──
    # Column A: field label
    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = thin_border

    for i, (code, rec) in enumerate(selected_records):
        col = i + 2  # Start from column B
        param_name = _format_field_value(rec.get("fields", {}).get(param_name_field, code))
        cell = ws.cell(row=1, column=col, value=param_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── Data Rows ──
    for row_idx, field_label in enumerate(METADATA_FIELDS):
        row = row_idx + 2

        # Column A: field label
        label_cell = ws.cell(row=row, column=1, value=field_label)
        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.alignment = left_align
        label_cell.border = thin_border

        for col_idx, (code, rec) in enumerate(selected_records):
            col = col_idx + 2
            fvals = rec.get("fields", {})

            # Try to find the value in Feishu fields
            value = fvals.get(field_label, "")

            # Special handling for date fields
            if isinstance(value, (int, float)) and value > 1000000000000:
                dt = _parse_timestamp(value)
                if dt:
                    value = dt.strftime("%Y-%m-%d %H:%M:%S")

            formatted = _format_field_value(value) if value else ""
            cell = ws.cell(row=row, column=col, value=formatted)
            cell.font = value_font
            cell.alignment = left_align
            cell.border = thin_border

    # ── Column Widths ──
    ws.column_dimensions["A"].width = 30.5
    for col_idx in range(len(selected_records)):
        col_letter = chr(ord("B") + col_idx)
        ws.column_dimensions[col_letter].width = 34.2

    # ── Row Heights ──
    ws.row_dimensions[1].height = 30
    for row_idx in range(len(METADATA_FIELDS)):
        ws.row_dimensions[row_idx + 2].height = 22

    # ── Freeze panes ──
    ws.freeze_panes = "B2"

    wb.save(output_path)
    return output_path


# ── CLI Interface ───────────────────────────────────────────────────

def _resolve_creds(args):
    """Resolve credentials with priority: CLI args > env vars > built-in defaults."""
    app_id = args.app_id or os.environ.get("FEISHU_APP_ID") or DEFAULT_APP_ID
    app_secret = args.app_secret or os.environ.get("FEISHU_APP_SECRET") or DEFAULT_APP_SECRET
    app_token = args.app_token or os.environ.get("FEISHU_APP_TOKEN") or DEFAULT_APP_TOKEN
    table_id = args.table_id or os.environ.get("FEISHU_TABLE_ID") or DEFAULT_TABLE_ID
    if args.table_url:
        parsed_token, parsed_table = parse_table_url(args.table_url)
        app_token = app_token if app_token != DEFAULT_APP_TOKEN else parsed_token or app_token
        table_id = table_id if table_id != DEFAULT_TABLE_ID else parsed_table or table_id
    return app_id, app_secret, app_token, table_id


def cmd_fetch(args):
    """Fetch command: discover fields and query records."""
    app_id, app_secret, app_token, table_id = _resolve_creds(args)

    missing = []
    if not app_id: missing.append("--app-id")
    if not app_secret: missing.append("--app-secret")
    if not app_token: missing.append("--app-token")
    if not table_id: missing.append("--table-id")
    if missing:
        print(f"错误: 缺少必要参数: {', '.join(missing)}")
        print("可通过命令行参数或环境变量提供 (FEISHU_APP_ID, FEISHU_APP_SECRET, etc.)")
        sys.exit(1)

    days = args.days or 7

    print(f"正在连接飞书 API...")
    token = get_tenant_access_token(app_id, app_secret)

    print(f"正在获取表格字段...")
    fields = list_fields(token, app_token, table_id)
    print(f"  获取到 {len(fields)} 个字段:")
    for f in fields:
        print(f"    - {f['field_name']} ({f['type_name']}) [{f['field_id']}]")

    print(f"\n正在获取表格记录...")
    records = list_records(token, app_token, table_id)
    print(f"  获取到 {len(records)} 条记录")

    print(f"\n正在按时间范围筛选 (近 {days} 天)...")
    filter_result = filter_records_by_time(records, days)

    # Display summary
    display_summary(fields, filter_result, days)

    # Output JSON if requested
    if args.output:
        output_data = {
            "fields": fields,
            "filter_result": {
                "since": filter_result["since"],
                "total_new": filter_result["total_new"],
                "total_updated": filter_result["total_updated"],
                "records": [
                    {
                        "record_id": r["record_id"],
                        "fields": r["fields"],
                        "created_time": r["created_time"].isoformat() if r["created_time"] else None,
                        "last_modified_time": r["last_modified_time"].isoformat() if r["last_modified_time"] else None,
                        "is_new": r in filter_result["new"],
                        "is_updated": r in filter_result["updated"],
                    }
                    for r in filter_result["all"]
                ],
            },
        }
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        print(f"\n完整数据已保存到: {args.output}")


def cmd_export(args):
    """Export command: generate Excel from selected parameters."""
    app_id, app_secret, app_token, table_id = _resolve_creds(args)

    missing = []
    if not app_id: missing.append("--app-id")
    if not app_secret: missing.append("--app-secret")
    if not app_token: missing.append("--app-token")
    if not table_id: missing.append("--table-id")
    if missing:
        print(f"错误: 缺少必要参数: {', '.join(missing)}")
        sys.exit(1)

    if not args.params:
        print("错误: 请指定要导出的参数编码 (--params)")
        sys.exit(1)

    selected_codes = [c.strip() for c in args.params.split(",")]
    output_path = args.output or "参数导出结果.xlsx"

    print(f"正在连接飞书 API...")
    token = get_tenant_access_token(app_id, app_secret)

    print(f"正在获取表格数据...")
    fields = list_fields(token, app_token, table_id)
    records = list_records(token, app_token, table_id)

    print(f"正在导出参数: {', '.join(selected_codes)}")
    export_to_excel(records, fields, selected_codes, output_path)

    print(f"\n✅ 导出成功: {output_path}")
    print(f"  共导出 {len(selected_codes)} 个参数，{len(METADATA_FIELDS)} 个元数据字段")


def main():
    parser = argparse.ArgumentParser(
        description="飞书多维表格监控与Excel导出工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", help="可用命令")

    # ── fetch ──
    fetch_parser = subparsers.add_parser("fetch", help="获取表格数据和变动记录")
    fetch_parser.add_argument("--app-id", help="飞书应用 App ID")
    fetch_parser.add_argument("--app-secret", help="飞书应用 App Secret")
    fetch_parser.add_argument("--app-token", help="飞书多维表格 App Token (wiki ID)")
    fetch_parser.add_argument("--table-id", help="飞书多维表格 Table ID")
    fetch_parser.add_argument("--table-url", help="飞书多维表格完整 URL (自动解析)")
    fetch_parser.add_argument("--days", type=int, default=7, help="监控天数 (默认: 7)")
    fetch_parser.add_argument("--output", help="JSON 输出文件路径")

    # ── export ──
    export_parser = subparsers.add_parser("export", help="导出指定参数到 Excel")
    export_parser.add_argument("--app-id", help="飞书应用 App ID")
    export_parser.add_argument("--app-secret", help="飞书应用 App Secret")
    export_parser.add_argument("--app-token", help="飞书多维表格 App Token")
    export_parser.add_argument("--table-id", help="飞书多维表格 Table ID")
    export_parser.add_argument("--table-url", help="飞书多维表格完整 URL (自动解析)")
    export_parser.add_argument("--params", required=True, help="要导出的参数编码，逗号分隔")
    export_parser.add_argument("--output", help="输出 Excel 文件路径")

    args = parser.parse_args()

    if args.command == "fetch":
        cmd_fetch(args)
    elif args.command == "export":
        cmd_export(args)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
